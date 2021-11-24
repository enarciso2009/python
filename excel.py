from openpyxl import load_workbook

arquivoTexto = open('file_d.txt','a')
caminho = 'c:/Everton/testes/testes_d.xlsx'

arquivo_excel=load_workbook(caminho)

planilha1 = arquivo_excel.active
planilha2 = arquivo_excel.active


print(arquivo_excel.sheetnames)

Matricula = planilha1['A1']
Nome = planilha1['B1']
Denominacao = planilha1['D1']
Horas = planilha1['G1']
Empresa = planilha1['H1']
PF = planilha1['I1']


max_linha = planilha1.max_row

print(f'{Matricula.value}, {Nome.value}, {Denominacao.value}, {Horas.value}, {Empresa.value}, {PF.value}')


#Horas.value = int(Horas.value)
print(Horas.value)
print(type(Horas.value))
print(PF.value)

cont = 0
for ind in range(2, max_linha):
    a = 'A' + str(ind)
    b = 'B' + str(ind)
    d = 'D' + str(ind)
    g = 'G' + str(ind)
    h = 'H' + str(ind)
    i = 'I' + str(ind)



    Matricula = planilha1[a]
    Nome = planilha1[b]
    Denominacao = planilha1[d]
    Horas = planilha1[g]
    Empresa = planilha1[h]
    PF = planilha1[i]

    print(Matricula.value, Nome.value, Denominacao.value, Horas.value, Empresa.value, PF.value)


    while True:
        #print(Matricula.value, Nome.value, Horas.value, Data.value, Ocorrencia.value, Codigo.value, Descricao.value)
        Horas.value = int(Horas.value)
        if Horas.value > 1380:
            arquivoTexto.write(str(Matricula.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Nome.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Denominacao.value))
            arquivoTexto.write(';')
            Horas.value -= 1380
            arquivoTexto.write(str(1380))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Empresa.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(PF.value))
            arquivoTexto.write('\n')
            print('aqui no > 1380')
            cont += 1
        elif Horas.value > 0 and Horas.value <= 1380:
            arquivoTexto.write(str(Matricula.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Nome.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Denominacao.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Horas.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(Empresa.value))
            arquivoTexto.write(';')
            arquivoTexto.write(str(PF.value))
            arquivoTexto.write('\n')
            print('esta aqui <= 1380')
            break
        elif Horas.value == 'None':
            print('none')
            break
        else:
            print('vai sair do loop')
            break

arquivoTexto.close()


'''
print('-' * 30)


var = 'A'+'5'
matr = planilha1[var]
print(matr.value)
print('_' * 30)



max_linha = planilha1.max_row
max_coluna = planilha1.max_column
new_sheet = wb.get_sheet_by_name('Sheet1')
dic = dict()

for n in range(1, max_linha):
    if Horas.value > 0:
        new_sheet.cell(row=1, column=0).value = Horas.value
    else:
        print(Horas.value)


for c in range(1, 5):
    perfil = planilha1['J' + str(c)]
    dic['perfil'] = perfil.value
    lista.append(dic)
    arquivo_excel.save("relatorio2.xlsx")


# Cria Novo workbook
wb = Workbook()
# Seleciona a active Sheet
ws1 = wb.active
# Rename it
ws1.title = 'my test'
# Escreve alguns dados
for col in range(1,5):
    for row in range(1,6):
        letter = get_column_letter(col)
        ws1[letter + str(row)] = letter + str(row)
# Cria nova sheet
ws2 = wb.create_sheet(title="Ok")
ws2["C1"] = "OK"
# Salva arquivo (Se não colocar o caminho complete, ele salva
# na mesma pasta do scritp.
wb.save('Text.xlsx')

'''




# str(self.cor))
